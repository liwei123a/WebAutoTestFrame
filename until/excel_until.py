# -*- coding:utf-8 -*-

import openpyxl

class ExcelUntil(object):
    def __init__(self, filename=None, sheetname=None):
        if filename is None:
            filename = 'config/casedata.xlsx'
        self.file = filename
        self.wb = openpyxl.load_workbook(self.file)
        self.sh = self.wb[sheetname]

    def get_table_rows(self):
        """
        获取总行数
        :return:
        """
        return self.sh.max_row

    def get_table_cols(self):
        """
        获取总列数
        :return:
        """
        return self.sh.max_column

    def get_cell_value(self, row, col):
        """
        获取单元格的值
        :param row: 行号
        :param col: 列号
        :return:
        """
        return self.sh.cell(row=int(row), column=int(col)).value

    def write_cell_value(self, row, col, value):
        """
        把数据写入单元格
        :param row:
        :param col:
        :param value:
        :return:
        """
        self.sh.cell(row=int(row), column=int(col), value=value)
        self.wb.save(self.file)

    def get_col_values(self, col):
        """
        获取某列所有值
        :param col:
        :return:
        """
        rows = self.get_table_rows()
        columndata = []
        for r in range(1, rows+1):
            cellvalue = self.get_cell_value(r, col)
            columndata.append(cellvalue)
        return columndata

    def get_row_values(self, row):
        """
        获取某行所有值
        :param row:
        :return:
        """
        columns = self.get_table_cols()
        rowdata = []
        for c in range(1,columns+1):
            cellvalue = self.get_cell_value(row, c)
            rowdata.append(cellvalue)
        return rowdata
